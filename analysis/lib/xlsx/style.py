from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils.cell import get_column_letter

# Guess at how many characters fit into a column width measurement
CHAR_PER_WIDTH_UNIT = 1.7

### Create named styles for formatting cells
font_bold = Font(bold=True)


alignment_left_wrap = Alignment(horizontal="left", wrap_text=True)
alignment_center_wrap = Alignment(horizontal="center", wrap_text=True)

default_header_border = Border(
    bottom=Side(border_style="medium", color="000000"),
)
default_cell_border = Border(
    bottom=Side(border_style="thin", color="AAAAAA"),
    left=Side(border_style="thin", color="DDDDDD"),
    right=Side(border_style="thin", color="DDDDDD"),
)

# Note: all cells are setup to wrap text
left_header_style = NamedStyle(
    name="Left Header Style",
    font=font_bold,
    alignment=alignment_left_wrap,
    border=default_header_border,
)

center_header_style = NamedStyle(
    name="Center Header Style",
    font=font_bold,
    alignment=alignment_center_wrap,
    border=default_header_border,
)

table_caption_style = NamedStyle(
    name="Table Header Style",
    font=font_bold,
    alignment=alignment_left_wrap,
)

good_condition_header_style = NamedStyle(
    name="Good Condition Header Style",
    alignment=alignment_center_wrap,
    border=Border(
        bottom=Side(border_style="thin", color="000000"),
    ),
    fill=PatternFill("solid", "EEEEEE"),
)


value_style = NamedStyle(
    name="Value Style",
    alignment=alignment_left_wrap,
    border=default_cell_border,
)

even_row_bg = PatternFill("solid", fgColor="00F6F6F6")

analysis_unit_divider = Border(
    bottom=Side(border_style="medium", color="AAAAAA"),
    left=Side(border_style="thin", color="DDDDDD"),
    right=Side(border_style="thin", color="DDDDDD"),
)

description_font = Font(color="999999")


def set_cell_styles(ws, breaks=None, area_columns=None, percent_columns=None):
    area_columns = area_columns or []
    percent_columns = percent_columns or []

    for col_idx, col in enumerate(ws.columns):
        col[0].style = center_header_style

        for i, cell in enumerate(col[1:]):
            cell.style = value_style
            value = cell.value
            is_int = isinstance(value, (float, int)) and int(value) == value

            if col_idx in area_columns:
                if is_int:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
            elif col_idx in percent_columns:
                if is_int:
                    cell.number_format = "0%"
                else:
                    cell.number_format = "0.00%"

            if i % 2 == 1:
                cell.fill = even_row_bg

    ws["A1"].style = left_header_style

    if breaks is not None:
        # add a stronger line between analysis units
        for col in ws.columns:
            for line in breaks:
                col[line].border = analysis_unit_divider


def set_column_widths(ws, widths):
    for i, width in enumerate(widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = width


def add_caption(ws, table_counter, caption):
    """Add a table caption followed by a blank line

    Parameters
    ----------
    ws : Worksheet
    table_counter : int
    caption : str
    """

    ws.insert_rows(idx=1, amount=2)

    cell = ws["A1"]
    cell.value = f"Table {table_counter}: {caption}"
    cell.style = table_caption_style

    end_col = get_column_letter(ws.max_column)
    ws.merge_cells(f"A1:{end_col}1")


def add_good_condition_row(ws, values_start_col, values_end_col, break_col):
    """Add header row with merged cells for not in good condition / in good condition

    Good conditions are only defined for indicators, which always have columns
    ordered greatest to least (good condition on the left.

    Parameters
    ----------
    ws : Worksheet
    values_start_col : int
        values start column index, 0-based
    values_end_col : int
        values end column index, 0-based
    break_col : int
        the column index of the first value after the break between good and not good condition, 0-based
    """
    row = 3

    ws.insert_rows(idx=row)

    start_col = get_column_letter(values_start_col + 1)
    end_col = get_column_letter(values_start_col + break_col)
    cell = ws[f"{start_col}{row}"]
    cell.value = "In good condition"
    cell.style = good_condition_header_style
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    start_col = get_column_letter(values_start_col + break_col + 1)
    end_col = get_column_letter(values_end_col)
    cell = ws[f"{start_col}{row}"]
    cell.value = "Not in good condition"
    cell.style = good_condition_header_style
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    for i in range(row, ws.max_row + row):
        cell = ws[f"{start_col}{i}"]
        cell.border = Border(
            left=Side(border_style="medium", color="666666"), bottom=cell.border.bottom, right=cell.border.right
        )
